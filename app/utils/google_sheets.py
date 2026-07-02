"""
Google Sheets integration helper.

Requires:
  - GOOGLE_SHEETS_SA_JSON env var: base64-encoded service account JSON key
    (Sheets API + Drive API must both be enabled on the Cloud project)
  - GOOGLE_DRIVE_FOLDER_ID env var: ID of a Shared Drive shared with the
    service account as Contributor. Files in a Shared Drive are owned by the
    drive, so they never count against individual or SA storage quotas.

One-time setup:
  1. Google Cloud Console → enable Sheets API + Drive API
  2. IAM & Admin → Service Accounts → Create → download JSON key
  3. Base64-encode: base64 -w0 service-account.json
  4. Add as GOOGLE_SHEETS_SA_JSON env var in Render
  5. Create a Shared Drive → add the service account EMAIL as Contributor
  6. Add the Shared Drive ID as GOOGLE_DRIVE_FOLDER_ID env var in Render
"""
from __future__ import annotations

import base64
import json
import os
import re
import threading
import time

import gspread
from gspread.exceptions import APIError
import openpyxl
from google.oauth2.service_account import Credentials
from google.auth.transport.requests import AuthorizedSession

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

_DRIVE_FILES_URL      = "https://www.googleapis.com/drive/v3/files"
_SHEETS_BASE_URL      = "https://sheets.googleapis.com/v4/spreadsheets"
_SEPARATOR_TEXT       = "No More Emails For Today."

# ── Sheet-read cache ──────────────────────────────────────────────────────────
# Multiple functions (read_leads, read_sent_with_dates, read_ab_stats) all call
# get_all_records on the same sheet.  During a sync, this would fire 2–3 reads
# per campaign within seconds and quickly exhaust the 60 reads/min quota.
#
# Strategy: cache the raw records for 45 seconds per (sheet_id, render_option).
# All reads within the same sync window share one API call.  Writes (date
# back-fills) call _invalidate_sheet_cache so the next read sees fresh data.
_records_cache: dict[str, tuple[float, list]] = {}
_records_cache_lock = threading.Lock()
_RECORDS_CACHE_TTL  = 45  # seconds


def _get_all_records(ws, sheet_id: str, **kwargs) -> list:
    """
    Fetch records from a worksheet with caching + 429-backoff.

    • Cache: returns the stored result if a read for this sheet + render option
      happened within _RECORDS_CACHE_TTL seconds.  Cuts API calls from 2–3 per
      campaign (sync + A/B) to 1 per campaign per 45-second window.

    • Retry: on HTTP 429 (quota exceeded) waits 10 s, 20 s, 40 s before the
      final re-raise so transient quota bursts self-heal rather than propagating
      an error to the user.

    Uses get_all_values + manual dict construction instead of get_all_records
    because gspread raises GSpreadException when the sheet header row contains
    duplicate column names (e.g. "First Name" appears in both the routing
    section before __divider__ and the prospect section after it).  We keep
    only the first occurrence of each duplicate key so callers that only care
    about unique columns (Recipient Email, Lead Status, etc.) are unaffected.
    """
    render = kwargs.get("value_render_option", "FORMATTED_VALUE")
    cache_key = f"{sheet_id}:{render}"
    now = time.time()

    with _records_cache_lock:
        entry = _records_cache.get(cache_key)
        if entry and (now - entry[0]) < _RECORDS_CACHE_TTL:
            return entry[1]

    # Cache miss — fetch from API with exponential backoff on 429
    wait_times = (10, 20, 40)
    for attempt in range(len(wait_times) + 1):
        try:
            all_values = ws.get_all_values(value_render_option=render)
            if not all_values:
                records: list = []
            else:
                raw_headers = all_values[0]
                # Deduplicate headers: keep first occurrence, mark repeats None.
                seen: set[str] = set()
                headers: list = []
                for h in raw_headers:
                    if h not in seen:
                        seen.add(h)
                        headers.append(h)
                    else:
                        headers.append(None)
                records = [
                    {h: v for h, v in zip(headers, row) if h is not None}
                    for row in all_values[1:]
                ]
            with _records_cache_lock:
                # Drop expired entries while we hold the lock so the cache
                # doesn't grow without bound across many sheets.
                fresh_now = time.time()
                for k in [k for k, (ts, _) in _records_cache.items()
                          if (fresh_now - ts) >= _RECORDS_CACHE_TTL]:
                    del _records_cache[k]
                _records_cache[cache_key] = (fresh_now, records)
            return records
        except APIError as exc:
            status = getattr(exc.response, "status_code", 0)
            if status == 429 and attempt < len(wait_times):
                time.sleep(wait_times[attempt])
            else:
                raise


def _invalidate_sheet_cache(sheet_id: str) -> None:
    """Remove all cached entries for a sheet (call after any write operation)."""
    prefix = f"{sheet_id}:"
    with _records_cache_lock:
        stale = [k for k in _records_cache if k.startswith(prefix)]
        for k in stale:
            del _records_cache[k]


# ── Auth helpers ──────────────────────────────────────────────────────────────

def is_configured() -> bool:
    """Return True if the Google Sheets env var is present."""
    return bool(os.environ.get("GOOGLE_SHEETS_SA_JSON", "").strip())


def _decode_sa_json() -> dict:
    """Decode the base64 service account JSON from the env var."""
    raw = os.environ.get("GOOGLE_SHEETS_SA_JSON", "").strip()
    if not raw:
        raise RuntimeError(
            "GOOGLE_SHEETS_SA_JSON env var is not set. "
            "Follow the setup instructions to configure Google Sheets integration."
        )
    try:
        return json.loads(base64.b64decode(raw + "=="))
    except Exception as exc:
        raise RuntimeError(f"Could not decode GOOGLE_SHEETS_SA_JSON: {exc}") from exc


# Module-level gspread client cache.
# Creating a new client for every call triggers a full OAuth token fetch each
# time (≈300–600 ms per call).  gspread's internal AuthorizedSession refreshes
# the token automatically when it nears expiry, so a long-lived client is safe.
_gc_cache: gspread.Client | None = None
_gc_lock  = threading.Lock()


def _client() -> gspread.Client:
    """Return a cached, authenticated gspread Client (gspread v6 native method).

    The client is created once per process and reused for all subsequent calls.
    gspread's AuthorizedSession refreshes the OAuth access token transparently
    when it expires, so no manual TTL management is needed.
    """
    global _gc_cache
    if _gc_cache is not None:
        return _gc_cache
    with _gc_lock:
        # Double-checked locking: another thread may have initialised it while
        # we were waiting for the lock.
        if _gc_cache is None:
            info = _decode_sa_json()
            _gc_cache = gspread.service_account_from_dict(info, scopes=SCOPES)
    return _gc_cache


# Module-level AuthorizedSession cache — mirrors _client().  Building a fresh
# session per call paid a full OAuth token fetch each time; a single sheet
# mark could create three sessions (one per CF helper).  AuthorizedSession
# refreshes its token automatically, so a long-lived instance is safe.
_session_cache: AuthorizedSession | None = None
_session_lock = threading.Lock()


def _authed_session() -> AuthorizedSession:
    """Return a cached AuthorizedSession for direct Sheets/Drive REST calls."""
    global _session_cache
    if _session_cache is not None:
        return _session_cache
    with _session_lock:
        if _session_cache is None:
            info = _decode_sa_json()
            creds = Credentials.from_service_account_info(info, scopes=SCOPES)
            _session_cache = AuthorizedSession(creds)
    return _session_cache


# ── Drive cleanup (admin utility) ─────────────────────────────────────────────

def cleanup_service_account_drive(older_than_days: int = 0) -> dict:
    """
    Delete spreadsheets owned by the service account.
    older_than_days=0 deletes ALL sheets; >0 deletes only older ones.
    Returns {"deleted": int, "errors": int}
    """
    session = _authed_session()
    q = "mimeType='application/vnd.google-apps.spreadsheet' and trashed=false"
    if older_than_days > 0:
        from datetime import datetime, timezone, timedelta
        cutoff = (datetime.now(timezone.utc) - timedelta(days=older_than_days)).strftime("%Y-%m-%dT%H:%M:%S")
        q += f" and createdTime < '{cutoff}'"

    files, page_token = [], None
    while True:
        params: dict = {"q": q, "fields": "nextPageToken,files(id,name)", "pageSize": 100}
        if page_token:
            params["pageToken"] = page_token
        resp = session.get(_DRIVE_FILES_URL, params=params)
        if not resp.ok:
            break
        data = resp.json()
        files.extend(data.get("files", []))
        page_token = data.get("nextPageToken")
        if not page_token:
            break

    deleted = errors = 0
    for f in files:
        r = session.delete(f"{_DRIVE_FILES_URL}/{f['id']}")
        if r.ok or r.status_code == 204:
            deleted += 1
        else:
            errors += 1
    return {"deleted": deleted, "errors": errors}


# ── Sheet creation ────────────────────────────────────────────────────────────

def _create_spreadsheet(title: str) -> tuple[str, str]:
    """
    Create a blank Google Spreadsheet inside the configured Shared Drive.

    Files in a Shared Drive are owned by the drive, not the service account,
    so they never count against individual or SA storage quotas.

    Returns (spreadsheet_id, spreadsheet_url).
    """
    drive_id = os.environ.get("GOOGLE_DRIVE_FOLDER_ID", "").strip()
    if not drive_id:
        raise RuntimeError(
            "GOOGLE_DRIVE_FOLDER_ID env var is not set. "
            "Set it to your Shared Drive ID."
        )

    session = _authed_session()
    resp = session.post(
        _DRIVE_FILES_URL,
        params={"supportsAllDrives": "true"},
        json={
            "name":     title,
            "mimeType": "application/vnd.google-apps.spreadsheet",
            "parents":  [drive_id],
        },
    )
    if not resp.ok:
        raise RuntimeError(
            f"Drive create failed (HTTP {resp.status_code}): {resp.text[:400]}"
        )
    sid = resp.json()["id"]
    url = f"https://docs.google.com/spreadsheets/d/{sid}/edit"
    return sid, url


# ── Colour / formatting helpers ───────────────────────────────────────────────

def _rgb(hex_str: str) -> dict:
    """Convert a 6-char hex colour string to a Sheets API colour object."""
    h = hex_str.lstrip("#")
    return {
        "red":   int(h[0:2], 16) / 255,
        "green": int(h[2:4], 16) / 255,
        "blue":  int(h[4:6], 16) / 255,
    }



def _apply_sheet_formatting(
    session: AuthorizedSession,
    spreadsheet_id: str,
    sheet_gid: int,
    all_rows: list[list[str]],   # header row + data rows (already str-converted)
) -> None:
    """
    Apply data-validation, conditional formatting, and per-row styles to the
    Outreach List sheet via a single Sheets API batchUpdate call.

    Replicates the Excel writer's formatting:
      - Send Status column    → checkbox (TRUE/FALSE)
      - Lead Status column    → dropdown (Lead / Reply / Unsubscribe)
      - First-of-sender rows  → pale yellow background (#FFFDE7), no bold
      - Separator rows        → merged, orange-tinted, bold brown text
      - Lead Status "Lead"    → green cell
      - Lead Status "Reply"   → orange cell
      - Lead Status "Unsubscribe" → red cell
      - Send Status = TRUE    → whole-row light green
    """
    if not all_rows:
        return

    headers   = all_rows[0]
    data_rows = all_rows[1:]
    n_cols    = len(headers)

    def col_idx(name: str) -> int:
        try:
            return headers.index(name)
        except ValueError:
            return -1

    send_status_col = col_idx("Send Status")
    lead_status_col = col_idx("Lead Status")
    sender_col      = col_idx("Sender Account")
    email_col       = col_idx("Recipient Email")
    div_col         = col_idx("__divider__")
    n_data          = len(data_rows)
    last_row_idx    = 1 + n_data   # exclusive end (0-based, row 0 = header)

    requests: list[dict] = []

    # ── Per-row styles (separator merge + yellow first-sender stripe) ─────
    prev_sender: str | None = None
    for ri, row in enumerate(data_rows):
        row_idx = ri + 1   # 0-based sheet row index

        # Detect separator row: only first cell has content = _SEPARATOR_TEXT
        is_sep = (
            row[0].strip() == _SEPARATOR_TEXT
            and all(v == "" for v in row[1:])
        )

        if is_sep:
            # Merge all cells in the separator row
            requests.append({"mergeCells": {
                "range": {
                    "sheetId":          sheet_gid,
                    "startRowIndex":    row_idx,
                    "endRowIndex":      row_idx + 1,
                    "startColumnIndex": 0,
                    "endColumnIndex":   n_cols,
                },
                "mergeType": "MERGE_ALL",
            }})
            # Style: warm orange background, bold brown text, centred
            requests.append({"repeatCell": {
                "range": {
                    "sheetId":          sheet_gid,
                    "startRowIndex":    row_idx,
                    "endRowIndex":      row_idx + 1,
                    "startColumnIndex": 0,
                    "endColumnIndex":   n_cols,
                },
                "cell": {"userEnteredFormat": {
                    "backgroundColor": _rgb("FFF3E0"),
                    "textFormat": {
                        "bold":            True,
                        "foregroundColor": _rgb("5D4037"),
                        "fontSize":        9,
                    },
                    "horizontalAlignment": "CENTER",
                    "verticalAlignment":   "MIDDLE",
                }},
                "fields": (
                    "userEnteredFormat(backgroundColor,textFormat,"
                    "horizontalAlignment,verticalAlignment)"
                ),
            }})
            prev_sender = None   # reset so first row of next day gets stripe

        else:
            # Yellow stripe for first row of each new sender block
            if sender_col >= 0 and sender_col < len(row):
                curr = row[sender_col]
                if curr != prev_sender:
                    requests.append({"repeatCell": {
                        "range": {
                            "sheetId":          sheet_gid,
                            "startRowIndex":    row_idx,
                            "endRowIndex":      row_idx + 1,
                            "startColumnIndex": 0,
                            "endColumnIndex":   n_cols,
                        },
                        "cell": {"userEnteredFormat": {
                            "backgroundColor": _rgb("FFFDE7"),
                        }},
                        "fields": "userEnteredFormat.backgroundColor",
                    }})
                    prev_sender = curr

    # ── Divider column (grey separator before prospect columns) ──────────
    # Matches the light-grey divider column written by excel_writer.py.
    # Header cell text is cleared; the whole column gets a grey fill and a
    # narrow pixel width so it reads as a visual boundary, not a data column.
    if div_col >= 0:
        # Header cell: dark grey fill, clear the "__divider__" text
        requests.append({"updateCells": {
            "rows": [{"values": [{"userEnteredFormat": {
                "backgroundColor": _rgb("BDBDBD"),
            }}]}],
            "fields": "userEnteredFormat.backgroundColor",
            "start": {
                "sheetId":     sheet_gid,
                "rowIndex":    0,
                "columnIndex": div_col,
            },
        }})
        requests.append({"updateCells": {
            "rows": [{"values": [{"userEnteredValue": {"stringValue": ""}}]}],
            "fields": "userEnteredValue",
            "start": {
                "sheetId":     sheet_gid,
                "rowIndex":    0,
                "columnIndex": div_col,
            },
        }})
        # Data rows: lighter grey fill
        if n_data > 0:
            requests.append({"repeatCell": {
                "range": {
                    "sheetId":          sheet_gid,
                    "startRowIndex":    1,
                    "endRowIndex":      last_row_idx,
                    "startColumnIndex": div_col,
                    "endColumnIndex":   div_col + 1,
                },
                "cell": {"userEnteredFormat": {
                    "backgroundColor": _rgb("EEEEEE"),
                }},
                "fields": "userEnteredFormat.backgroundColor",
            }})
        # Narrow column width (~20 px ≈ 3-char wide visual separator)
        requests.append({"updateDimensionProperties": {
            "range": {
                "sheetId":    sheet_gid,
                "dimension":  "COLUMNS",
                "startIndex": div_col,
                "endIndex":   div_col + 1,
            },
            "properties": {"pixelSize": 20},
            "fields": "pixelSize",
        }})

    # ── Data validation: Send Status = checkbox ───────────────────────────
    if send_status_col >= 0:
        requests.append({"setDataValidation": {
            "range": {
                "sheetId":          sheet_gid,
                "startRowIndex":    1,
                "endRowIndex":      last_row_idx,
                "startColumnIndex": send_status_col,
                "endColumnIndex":   send_status_col + 1,
            },
            "rule": {"condition": {"type": "BOOLEAN"}, "showCustomUi": True},
        }})

    # ── Data validation: Chaser Sent? = checkbox ──────────────────────────
    chaser_col = col_idx("Chaser Sent?")
    if chaser_col >= 0:
        requests.append({"setDataValidation": {
            "range": {
                "sheetId":          sheet_gid,
                "startRowIndex":    1,
                "endRowIndex":      last_row_idx,
                "startColumnIndex": chaser_col,
                "endColumnIndex":   chaser_col + 1,
            },
            "rule": {"condition": {"type": "BOOLEAN"}, "showCustomUi": True},
        }})

    # ── Data validation: Lead Status = dropdown ───────────────────────────
    if lead_status_col >= 0:
        requests.append({"setDataValidation": {
            "range": {
                "sheetId":          sheet_gid,
                "startRowIndex":    1,
                "endRowIndex":      last_row_idx,
                "startColumnIndex": lead_status_col,
                "endColumnIndex":   lead_status_col + 1,
            },
            "rule": {
                "condition": {
                    "type":   "ONE_OF_LIST",
                    "values": [
                        {"userEnteredValue": "Lead"},
                        {"userEnteredValue": "Reply"},
                        {"userEnteredValue": "Interested"},
                        {"userEnteredValue": "Unsubscribe"},
                    ],
                },
                "showCustomUi": True,
                "strict":       False,
            },
        }})

    # ── Conditional formatting ─────────────────────────────────────────────
    # Rules are added with index=0. Each insertion shifts previous rules down,
    # so the LAST rule added ends up at the highest priority (index 0).
    #
    # Priority order (low → high):
    #   1. Domain-grey  whole-row grey  (lowest  — overridden by send & lead)
    #   2. Send Status  whole-row blue  (middle)
    #   3. Lead Status  cell colours    (highest — narrow range, always wins)

    # 1. Send Status checked → whole-row light BLUE (lowest priority — added first).
    # Domain-grey is added after (higher priority) so blocked-domain rows stay
    # grey even when they have been marked as sent.
    if send_status_col >= 0:
        ss_letter = _col_letter(send_status_col + 1)
        requests.append({"addConditionalFormatRule": {
            "rule": {
                "ranges": [{
                    "sheetId":          sheet_gid,
                    "startRowIndex":    1,
                    "endRowIndex":      last_row_idx,
                    "startColumnIndex": 0,
                    "endColumnIndex":   n_cols,
                }],
                "booleanRule": {
                    "condition": {
                        "type":   "CUSTOM_FORMULA",
                        "values": [{"userEnteredValue": f"=${ss_letter}2=TRUE"}],
                    },
                    "format": {"backgroundColor": _rgb("E3F2FD")},  # light blue
                },
            },
            "index": 0,
        }})

    # 2. Domain-grey — higher priority than sent-blue so blocked-domain rows stay
    # grey even when the prospect has been marked as sent (blue).
    # Formula: if this row is NOT already Lead AND its email domain appears on any
    # Lead row in the sheet, grey the whole row.
    if lead_status_col >= 0 and email_col >= 0:
        ls_letter = _col_letter(lead_status_col + 1)
        em_letter = _col_letter(email_col + 1)
        domain_formula = (
            f'=AND(${ls_letter}2<>"Lead",'
            f'ISNUMBER(FIND("@",${em_letter}2)),'
            f'COUNTIFS(${ls_letter}$2:${ls_letter}${last_row_idx},"Lead",'
            f'${em_letter}$2:${em_letter}${last_row_idx},'
            f'"*@"&RIGHT(${em_letter}2,LEN(${em_letter}2)-FIND("@",${em_letter}2)))>0)'
        )
        requests.append({"addConditionalFormatRule": {
            "rule": {
                "ranges": [{
                    "sheetId":          sheet_gid,
                    "startRowIndex":    1,
                    "endRowIndex":      last_row_idx,
                    "startColumnIndex": 0,
                    "endColumnIndex":   n_cols,
                }],
                "booleanRule": {
                    "condition": {
                        "type":   "CUSTOM_FORMULA",
                        "values": [{"userEnteredValue": domain_formula}],
                    },
                    "format": {"backgroundColor": _rgb("E0E0E0")},
                },
            },
            "index": 0,
        }})

    # 4. Lead Status cell colours (LS column only — added before lead-row-grey so
    # lead-row-grey ends up with higher priority for non-LS cells).
    if lead_status_col >= 0:
        ls_range = {
            "sheetId":          sheet_gid,
            "startRowIndex":    1,
            "endRowIndex":      last_row_idx,
            "startColumnIndex": lead_status_col,
            "endColumnIndex":   lead_status_col + 1,
        }
        for value, bg_hex, fg_hex in [
            ("Lead",        "66BB6A", "FFFFFF"),   # vivid green / white
            ("Reply",       "FFA726", "FFFFFF"),   # vivid amber / white
            ("Interested",  "AB47BC", "FFFFFF"),   # vivid purple / white
            ("Unsubscribe", "EF5350", "FFFFFF"),   # vivid red / white
        ]:
            requests.append({"addConditionalFormatRule": {
                "rule": {
                    "ranges": [ls_range],
                    "booleanRule": {
                        "condition": {
                            "type":   "TEXT_EQ",
                            "values": [{"userEnteredValue": value}],
                        },
                        "format": {
                            "backgroundColor": _rgb(bg_hex),
                            "textFormat": {
                                "foregroundColor": _rgb(fg_hex),
                                "bold": True,
                            },
                        },
                    },
                },
                "index": 0,
            }})

    # 4.5. Any Lead Status value → whole-row grey EXCEPT the LS column.
    # Fires for Lead, Reply, Interested AND Unsubscribe — each greys its own
    # row only.  (Domain-wide greying of other same-domain rows remains
    # Lead-only via the domain-grey rule above.)
    # Added LAST so it ends up at index 0 (highest priority) after all rules are
    # processed.  This ensures the grey overrides the sent-blue fill even when a
    # row has both Send Status = TRUE and a Lead Status value.
    # The LS column is excluded from the ranges so the coloured status cell
    # (rule 4) is never covered — only non-LS cells go grey.
    if lead_status_col >= 0:
        ls_letter    = _col_letter(lead_status_col + 1)
        lead_formula = (
            f'=OR(${ls_letter}2="Lead",${ls_letter}2="Reply",'
            f'${ls_letter}2="Interested",${ls_letter}2="Unsubscribe")'
        )
        lead_ranges: list[dict] = []
        if lead_status_col > 0:
            lead_ranges.append({
                "sheetId":          sheet_gid,
                "startRowIndex":    1,
                "endRowIndex":      last_row_idx,
                "startColumnIndex": 0,
                "endColumnIndex":   lead_status_col,   # up to (not including) LS col
            })
        if lead_status_col + 1 < n_cols:
            lead_ranges.append({
                "sheetId":          sheet_gid,
                "startRowIndex":    1,
                "endRowIndex":      last_row_idx,
                "startColumnIndex": lead_status_col + 1,  # one after LS col
                "endColumnIndex":   n_cols,
            })
        if lead_ranges:
            requests.append({"addConditionalFormatRule": {
                "rule": {
                    "ranges": lead_ranges,
                    "booleanRule": {
                        "condition": {
                            "type":   "CUSTOM_FORMULA",
                            "values": [{"userEnteredValue": lead_formula}],
                        },
                        "format": {"backgroundColor": _rgb("E0E0E0")},
                    },
                },
                "index": 0,
            }})

    # 5. First-of-sender yellow stripe (absolute highest priority — added last).
    # The per-row yellow applied as direct cell formatting during sheet creation
    # is overridden by CF rules (sent-blue, domain-grey, lead-row-grey).  This
    # CF rule re-applies yellow to the Sender Account cell only for the first
    # row of each sender group, so the yellow persists regardless of sent/lead
    # status.  Scoped to the Sender Account column only so the rest of the row
    # still reflects the correct sent/lead colour.
    if sender_col >= 0:
        sa_letter     = _col_letter(sender_col + 1)
        sender_formula = f'=AND(${sa_letter}2<>"",${sa_letter}2<>${sa_letter}1)'
        requests.append({"addConditionalFormatRule": {
            "rule": {
                "ranges": [{
                    "sheetId":          sheet_gid,
                    "startRowIndex":    1,
                    "endRowIndex":      last_row_idx,
                    "startColumnIndex": sender_col,
                    "endColumnIndex":   sender_col + 1,   # Sender Account column only
                }],
                "booleanRule": {
                    "condition": {
                        "type":   "CUSTOM_FORMULA",
                        "values": [{"userEnteredValue": sender_formula}],
                    },
                    "format": {"backgroundColor": _rgb("FFFDE7")},  # pale yellow
                },
            },
            "index": 0,
        }})

    if not requests:
        return

    resp = session.post(
        f"{_SHEETS_BASE_URL}/{spreadsheet_id}:batchUpdate",
        json={"requests": requests},
    )
    if not resp.ok:
        print(f"[google_sheets] batchUpdate FAILED {resp.status_code}: {resp.text[:600]}")
    else:
        print(f"[google_sheets] batchUpdate OK — {len(requests)} requests applied")


# ── Public: create outreach sheet ─────────────────────────────────────────────

def create_outreach_sheet(title: str, xlsx_path: str) -> dict:
    """
    Read the 'Outreach List' sheet from an openpyxl xlsx file
    (generated by mail_merge/utils/excel_writer.py), write all rows to a new
    Google Sheet in the configured Shared Drive, apply Excel-equivalent
    formatting (dropdowns, conditional formatting, sender-change stripes,
    separator rows), freeze the header, and share as "anyone with link can edit".

    Returns: {"sheet_id": str, "sheet_url": str, "title": str}
    """
    # ── Read Excel ────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_name = "Outreach List" if "Outreach List" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    raw_rows = [[cell.value for cell in row] for row in ws.iter_rows()]
    if not raw_rows:
        raise ValueError("The merge output file is empty.")

    str_rows = [[str(v) if v is not None else "" for v in row] for row in raw_rows]

    headers   = str_rows[0]
    data_rows = str_rows[1:]
    all_rows  = str_rows

    # Detect whether this merge actually has chaser content (vs. placeholder cols)
    _cb_idx   = headers.index("Chaser Body") if "Chaser Body" in headers else -1
    has_chaser = _cb_idx >= 0 and any(
        row[_cb_idx] for row in data_rows if _cb_idx < len(row)
    )

    # Log detected column positions for debugging
    _div_idx = headers.index("__divider__") if "__divider__" in headers else -1
    _ls_idx  = headers.index("Lead Status") if "Lead Status" in headers else -1
    _em_idx  = headers.index("Recipient Email") if "Recipient Email" in headers else -1
    print(
        f"[google_sheets] columns — __divider__:{_div_idx}  "
        f"LeadStatus:{_ls_idx}  RecipientEmail:{_em_idx}  total:{len(headers)}"
    )

    # ── Create spreadsheet in Shared Drive ───────────────────────────────
    file_id, sheet_url = _create_spreadsheet(title)

    # ── Write data via gspread ───────────────────────────────────────────
    gc     = _client()
    sh     = gc.open_by_key(file_id)
    gsheet = sh.sheet1
    gsheet.update_title("Outreach List")
    gsheet.update(all_rows, "A1")

    # Bold header + freeze
    gsheet.format(
        f"A1:{_col_letter(len(headers))}1",
        {"textFormat": {"bold": True}},
    )
    gsheet.freeze(rows=1)

    # ── Apply full formatting via batchUpdate ─────────────────────────────
    try:
        session = _authed_session()
        print(f"[google_sheets] applying formatting (sheet_gid={gsheet.id}) …")
        _apply_sheet_formatting(session, file_id, gsheet.id, all_rows)
        print("[google_sheets] formatting applied OK")
    except Exception as fmt_err:
        print(f"[google_sheets] formatting FAILED: {fmt_err}")

    # ── Share: anyone with link can edit ─────────────────────────────────
    sh.share("", perm_type="anyone", role="writer")

    # ── Hide chaser columns when there are no chasers ────────────────────
    if not has_chaser:
        _CHASER_COLS = ['Chaser Send Time', 'Chaser Body', 'Chaser Sent?', 'Chaser Date']
        hide_requests = []
        for col_name in _CHASER_COLS:
            if col_name in headers:
                col_0 = headers.index(col_name)
                hide_requests.append({
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId":    gsheet.id,
                            "dimension":  "COLUMNS",
                            "startIndex": col_0,
                            "endIndex":   col_0 + 1,
                        },
                        "properties": {"hiddenByUser": True},
                        "fields": "hiddenByUser",
                    }
                })
        if hide_requests:
            try:
                session = _authed_session()
                session.post(
                    f"{_SHEETS_BASE_URL}/{file_id}:batchUpdate",
                    json={"requests": hide_requests},
                )
                print(f"[google_sheets] hid {len(hide_requests)} chaser column(s)")
            except Exception as hide_err:
                print(f"[google_sheets] column hide FAILED (non-fatal): {hide_err}")

    # ── Stats tab ─────────────────────────────────────────────────────────
    try:
        _add_stats_sheet(sh, headers, has_chaser)
    except Exception as stats_err:
        print(f"[google_sheets] Stats sheet FAILED (non-fatal): {stats_err}")

    return {
        "sheet_id":  file_id,
        "sheet_url": sheet_url,
        "title":     title,
    }


# ── Public: read sheet data ───────────────────────────────────────────────────

def extract_sheet_id(url_or_id: str) -> str:
    """
    Accept either a full Google Sheets URL or a bare spreadsheet ID
    and return just the ID component.
    """
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9\-_]+)", url_or_id)
    return m.group(1) if m else url_or_id.strip()


def _is_sent(value) -> bool:
    """Return True if a Send Status cell value counts as 'sent'.

    Handles all representations gspread may return:
      • Python bool True  (UNFORMATTED_VALUE render)
      • String "TRUE"     (FORMATTED_VALUE render)
      • String "Sent"     (legacy dropdown before checkbox migration)
      • Integer 1         (edge case numeric true)
    """
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value == 1
    return str(value).strip().upper() in ("TRUE", "SENT")


def _to_iso_date(value) -> str:
    """Normalise a sheet date cell to YYYY-MM-DD, or "" if unparseable.

    Cells the app writes are RAW ISO strings, but hand-edited cells can come
    back as Google serial numbers — UNFORMATTED_VALUE renders real date cells
    as days-since-1899-12-30 (e.g. 46100 instead of "2026-03-19").  Passing a
    serial through to Supabase would fail the whole insert chunk, silently
    dropping those sends from the stats, so anything unparseable returns ""
    and gets back-filled (and rewritten to the sheet) as today during sync.
    """
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, (int, float)):
        from datetime import date as _d, timedelta as _td
        serial = int(value)
        if 1 <= serial < 200000:   # plausible Sheets serial range
            try:
                return (_d(1899, 12, 30) + _td(days=serial)).isoformat()
            except Exception:
                return ""
        return ""
    m = re.match(r"^(\d{4}-\d{2}-\d{2})", str(value).strip())
    return m.group(1) if m else ""


def read_sent_with_dates(sheet_id: str) -> list[dict]:
    """
    Return [{email, sent_date, row_num}] for all rows where Send Status is ticked.

    sent_date is the value from the 'Sent Date' column (ISO date string, may be
    "" if the column doesn't exist yet or has not been populated for that row).
    row_num is 1-based sheet row number (row 1 = header, row 2 = first data row).
    """
    gc = _client()
    sh = gc.open_by_key(sheet_id)
    ws = sh.sheet1
    records = _get_all_records(ws, sheet_id, value_render_option="UNFORMATTED_VALUE")

    results: list[dict] = []
    for i, r in enumerate(records, start=2):   # row 2 = first data row
        email = str(r.get("Recipient Email", "")).strip().lower()
        if not _is_sent(r.get("Send Status", "")) or not email or "@" not in email:
            continue
        sent_date    = _to_iso_date(r.get("Sent Date", ""))
        chaser_sent  = _is_sent(r.get("Chaser Sent?", ""))
        chaser_date  = _to_iso_date(r.get("Chaser Date", ""))
        results.append({"email": email, "sent_date": sent_date,
                         "row_num": i, "chaser_sent": chaser_sent,
                         "chaser_date": chaser_date})
    return results


def _write_date_column(
    sheet_id: str,
    row_date_pairs: list[tuple[int, str]],
    col_name: str,
) -> None:
    """
    Batch-write (row_num, date_str) pairs to a named date column.

    If the column doesn't exist in the header row it is appended automatically.
    row_num values are 1-based sheet row numbers (1 = header).
    """
    if not row_date_pairs:
        return

    gc = _client()
    sh = gc.open_by_key(sheet_id)
    ws = sh.sheet1

    # Prefer the already-cached records (populated by read_sent_with_dates
    # moments earlier during the same sync) over a fresh row_values(1) call.
    # This avoids an extra uncached API round-trip that has no quota-retry
    # protection and often pushes syncs over the 60-reads/min quota when
    # multiple campaigns sync in quick succession.
    records = _get_all_records(ws, sheet_id, value_render_option="UNFORMATTED_VALUE")
    headers: list[str] = list(records[0].keys()) if records else ws.row_values(1)

    if col_name not in headers:
        # Column absent — append it.  Use the raw row to count all columns
        # (including trailing empties that get_all_records may trim).
        raw_headers = ws.row_values(1)
        new_col_idx = len(raw_headers) + 1       # 1-based
        ws.update_cell(1, new_col_idx, col_name)
        date_col = new_col_idx
        _invalidate_sheet_cache(sheet_id)        # header changed
    else:
        date_col = headers.index(col_name) + 1   # 1-based

    updates = [
        {"range": f"{_col_letter(date_col)}{row_num}", "values": [[date_str]]}
        for row_num, date_str in row_date_pairs
    ]
    if not updates:
        return

    # Retry once on quota errors (429) — the read path already consumed some
    # quota, and large sheets can tip the burst budget on the write.
    wait_times = (10, 20)
    for attempt in range(len(wait_times) + 1):
        try:
            ws.batch_update(updates)
            break
        except APIError as exc:
            status = getattr(exc.response, "status_code", 0)
            if status == 429 and attempt < len(wait_times):
                time.sleep(wait_times[attempt])
            else:
                raise

    # Invalidate cache so the next read sees the freshly written dates.
    _invalidate_sheet_cache(sheet_id)


def write_sent_dates(sheet_id: str, row_date_pairs: list[tuple[int, str]]) -> None:
    """Batch-write to the 'Sent Date' helper column (auto-creates if missing)."""
    _write_date_column(sheet_id, row_date_pairs, "Sent Date")


def write_chaser_dates(sheet_id: str, row_date_pairs: list[tuple[int, str]]) -> None:
    """Batch-write to the 'Chaser Date' helper column (auto-creates if missing).

    Called during sync for every row where 'Chaser Sent?' is TRUE but 'Chaser Date'
    is still empty — stamps today's date so chaser sends are date-trackable for
    the Today/Week/Month stats.
    """
    _write_date_column(sheet_id, row_date_pairs, "Chaser Date")


def read_ab_stats(sheet_id: str) -> list[dict]:
    """
    Group sheet rows by 'Template Variant' and count Lead Status outcomes.

    Only rows with a valid email address and a non-empty Template Variant
    are counted; separator rows are skipped.

    Returns a list of dicts sorted by positive responses (lead + interested)
    descending, then by reply count:

    [
      {
        "variant":    "S2/B2",
        "total":      45,        # total prospects allocated to this variant
        "lead":       10,
        "interested": 2,
        "reply":      5,
        "unsubscribe": 0,
        "positive":   12,        # lead + interested
      },
      ...
    ]
    """
    from collections import defaultdict

    gc     = _client()
    sh     = gc.open_by_key(sheet_id)
    gsheet = sh.sheet1
    records = _get_all_records(gsheet, sheet_id, value_render_option="UNFORMATTED_VALUE")

    stats: dict = defaultdict(
        lambda: {"total": 0, "lead": 0, "interested": 0, "reply": 0, "unsubscribe": 0}
    )

    for r in records:
        email   = str(r.get("Recipient Email", "")).strip()
        variant = str(r.get("A/B Variant", "")).strip()
        if not email or "@" not in email or not variant:
            continue  # separator rows or rows without variant assignment

        status = str(r.get("Lead Status", "")).strip()
        stats[variant]["total"] += 1
        if status == "Lead":
            stats[variant]["lead"] += 1
        elif status == "Interested":
            stats[variant]["interested"] += 1
        elif status == "Reply":
            stats[variant]["reply"] += 1
        elif status == "Unsubscribe":
            stats[variant]["unsubscribe"] += 1

    result = []
    for variant, s in stats.items():
        positive = s["lead"] + s["interested"]
        result.append({
            "variant":     variant,
            "total":       s["total"],
            "lead":        s["lead"],
            "interested":  s["interested"],
            "reply":       s["reply"],
            "unsubscribe": s["unsubscribe"],
            "positive":    positive,
        })

    result.sort(key=lambda x: (x["positive"], x["reply"]), reverse=True)
    return result


def read_leads(sheet_id: str) -> list[dict]:
    """
    Return {"email", "status"} pairs where Lead Status is "Lead" or "Unsubscribe".
    Rows with status "Reply" or blank are ignored.

    Uses UNFORMATTED_VALUE so the result is shared with read_sent_with_dates
    in the cache — sync_campaign_core calls both, halving API calls per campaign.
    """
    gc = _client()
    sh = gc.open_by_key(sheet_id)
    gsheet = sh.sheet1
    records = _get_all_records(gsheet, sheet_id, value_render_option="UNFORMATTED_VALUE")

    results: list[dict] = []
    for record in records:
        status = str(record.get("Lead Status", "")).strip()
        email  = str(record.get("Recipient Email", "")).strip().lower()
        if status in ("Lead", "Unsubscribe", "Interested", "Reply") and email and "@" in email:
            results.append({"email": email, "status": status})
    return results


# ── Lead-grey CF rule helper ──────────────────────────────────────────────────

def _existing_cf_formulas(sheet_id: str, ws_id: int) -> set[str] | None:
    """
    Return the set of CUSTOM_FORMULA strings already present as conditional-
    format rules on the given worksheet, or None if the lookup failed.

    Used by mark_email_in_sheet so the retroactive CF helpers below only add
    their rule when it is genuinely missing.  Without this check every mark
    re-added 2–3 identical rules at index 0, so sheets steadily accumulated
    duplicate rules (slower rendering, cluttered rules list).

    Callers should treat None as "unknown" and fall back to adding the rule
    (the old, always-add behaviour) — a duplicate rule is cosmetic, a missing
    rule breaks the grey-out logic.
    """
    try:
        session = _authed_session()
        resp = session.get(
            f"{_SHEETS_BASE_URL}/{sheet_id}",
            params={"fields": "sheets(properties(sheetId),conditionalFormats)"},
        )
        if not resp.ok:
            return None
        formulas: set[str] = set()
        for sheet in resp.json().get("sheets", []):
            if sheet.get("properties", {}).get("sheetId") != ws_id:
                continue
            for rule in sheet.get("conditionalFormats") or []:
                cond = (rule.get("booleanRule") or {}).get("condition") or {}
                if cond.get("type") == "CUSTOM_FORMULA":
                    for v in cond.get("values") or []:
                        f = v.get("userEnteredValue")
                        if f:
                            formulas.add(f)
        return formulas
    except Exception:
        return None


def _apply_domain_grey_cf(
    sheet_id:   str,
    ws_id:      int,   # Google Sheets GID of the worksheet
    ls_col_idx: int,   # 1-based column index of "Lead Status"
    em_col_idx: int,   # 1-based column index of "Recipient Email"
    n_cols:     int,
    existing_formulas: set[str] | None = None,
) -> None:
    """
    Add the domain-grey CF rule to an existing sheet via direct REST call.

    Greys ALL rows that share the email domain of any Lead row (excluding the
    Lead row itself, which is handled by _apply_lead_grey_cf).

    Added at index=0 (highest priority) so the grey overrides the sent-blue
    fill — ensuring blocked-domain prospects appear grey even after being sent.

    Called from mark_email_in_sheet for both Email+Lead and Domain+Lead DNC
    entries so old sheets (created before the rule was baked into
    _apply_sheet_formatting) get the rule retroactively.
    """
    ls_letter = _col_letter(ls_col_idx)
    em_letter = _col_letter(em_col_idx)
    row_count = 3000
    formula = (
        f'=AND(${ls_letter}2<>"Lead",'
        f'ISNUMBER(FIND("@",${em_letter}2)),'
        f'COUNTIFS(${ls_letter}$2:${ls_letter}{row_count},"Lead",'
        f'${em_letter}$2:${em_letter}{row_count},'
        f'"*@"&RIGHT(${em_letter}2,LEN(${em_letter}2)-FIND("@",${em_letter}2)))>0)'
    )
    if existing_formulas is not None and formula in existing_formulas:
        return   # rule already on the sheet — don't add a duplicate
    session = _authed_session()
    resp = session.post(
        f"{_SHEETS_BASE_URL}/{sheet_id}:batchUpdate",
        json={"requests": [{
            "addConditionalFormatRule": {
                "rule": {
                    "ranges": [{
                        "sheetId":          ws_id,
                        "startRowIndex":    1,
                        "endRowIndex":      row_count,
                        "startColumnIndex": 0,
                        "endColumnIndex":   n_cols,
                    }],
                    "booleanRule": {
                        "condition": {
                            "type":   "CUSTOM_FORMULA",
                            "values": [{"userEnteredValue": formula}],
                        },
                        "format": {"backgroundColor": _rgb("E0E0E0")},
                    },
                },
                "index": 0,
            }
        }]},
    )
    if not resp.ok:
        print(f"[google_sheets] _apply_domain_grey_cf FAILED {resp.status_code}: {resp.text[:300]}")


def _apply_sender_stripe_cf(
    sheet_id:       str,
    ws_id:          int,   # Google Sheets GID
    sender_col_idx: int,   # 1-based column index of "Sender Account"
    existing_formulas: set[str] | None = None,
) -> None:
    """
    Add the first-of-sender yellow stripe CF rule to an existing sheet.

    Scoped to the Sender Account column only — applies pale yellow (#FFFDE7)
    whenever the Sender Account value differs from the row above (= first row
    of a new sender group).  Added at index=0 (highest priority) so it
    persists over sent-blue and lead/domain-grey fills.
    """
    sa_0      = sender_col_idx - 1         # 0-based
    sa_letter = _col_letter(sender_col_idx)
    formula   = f'=AND(${sa_letter}2<>"",${sa_letter}2<>${sa_letter}1)'

    if existing_formulas is not None and formula in existing_formulas:
        return   # rule already on the sheet — don't add a duplicate
    session = _authed_session()
    resp = session.post(
        f"{_SHEETS_BASE_URL}/{sheet_id}:batchUpdate",
        json={"requests": [{
            "addConditionalFormatRule": {
                "rule": {
                    "ranges": [{
                        "sheetId":          ws_id,
                        "startRowIndex":    1,
                        "endRowIndex":      3000,
                        "startColumnIndex": sa_0,
                        "endColumnIndex":   sa_0 + 1,
                    }],
                    "booleanRule": {
                        "condition": {
                            "type":   "CUSTOM_FORMULA",
                            "values": [{"userEnteredValue": formula}],
                        },
                        "format": {"backgroundColor": _rgb("FFFDE7")},
                    },
                },
                "index": 0,
            }
        }]},
    )
    if not resp.ok:
        print(f"[google_sheets] _apply_sender_stripe_cf FAILED {resp.status_code}: {resp.text[:300]}")


def _apply_lead_grey_cf(
    sheet_id:  str,
    ws_id:     int,   # Google Sheets GID of the worksheet
    ls_col_idx: int,  # 1-based column index of "Lead Status"
    n_cols:    int,
    existing_formulas: set[str] | None = None,
) -> None:
    """
    Add a lead-row-grey CF rule to an existing sheet via a direct REST call.

    Called from mark_email_in_sheet for sheets created before the rule was
    added to _apply_sheet_formatting, and as a belt-and-suspenders measure
    for all lead markings to ensure the rule is always present.

    Uses the raw AuthorizedSession (same as _apply_sheet_formatting) rather
    than the gspread Spreadsheet.batch_update() wrapper, which avoids any
    potential gspread version mismatch and gives us error logging.

    The rule is inserted at index 0 (highest priority in Google Sheets API)
    so it overrides both the sent-blue fill AND the domain-grey fill for rows
    with any Lead Status value (Lead / Reply / Interested / Unsubscribe) —
    each status greys its own row only.  The LS column is excluded from the
    ranges so the coloured status cell is never covered.
    """
    ls_0      = ls_col_idx - 1          # 0-based column index
    ls_letter = _col_letter(ls_col_idx)
    formula   = (
        f'=OR(${ls_letter}2="Lead",${ls_letter}2="Reply",'
        f'${ls_letter}2="Interested",${ls_letter}2="Unsubscribe")'
    )
    if existing_formulas is not None and formula in existing_formulas:
        return   # rule already on the sheet — don't add a duplicate
    row_count = 3000   # generous upper bound — covers any realistic sheet

    ranges: list[dict] = []
    if ls_0 > 0:
        ranges.append({
            "sheetId":          ws_id,
            "startRowIndex":    1,
            "endRowIndex":      row_count,
            "startColumnIndex": 0,
            "endColumnIndex":   ls_0,
        })
    if ls_0 + 1 < n_cols:
        ranges.append({
            "sheetId":          ws_id,
            "startRowIndex":    1,
            "endRowIndex":      row_count,
            "startColumnIndex": ls_0 + 1,
            "endColumnIndex":   n_cols,
        })
    if not ranges:
        return

    session = _authed_session()
    resp = session.post(
        f"{_SHEETS_BASE_URL}/{sheet_id}:batchUpdate",
        json={"requests": [{
            "addConditionalFormatRule": {
                "rule": {
                    "ranges": ranges,
                    "booleanRule": {
                        "condition": {
                            "type":   "CUSTOM_FORMULA",
                            "values": [{"userEnteredValue": formula}],
                        },
                        "format": {"backgroundColor": _rgb("E0E0E0")},
                    },
                },
                "index": 0,
            }
        }]},
    )
    if not resp.ok:
        print(f"[google_sheets] _apply_lead_grey_cf FAILED {resp.status_code}: {resp.text[:300]}")


# ── DNC reason → Lead Status mapping ─────────────────────────────────────────

_REASON_TO_LEAD_STATUS: dict[str, str] = {
    "lead":       "Lead",
    "interested": "Interested",
    "reply":      "Reply",
    "opt_out":    "Unsubscribe",
    "manual":     "Unsubscribe",
}


def mark_email_in_sheet(sheet_id: str, email_or_domain: str, reason: str = "manual") -> int:
    """
    Find the row(s) in the sheet where 'Recipient Email' **exactly** matches
    *email_or_domain* and set 'Lead Status' to the value for *reason*.

    Always uses exact email matching — a bare domain (no "@") returns 0
    immediately because it cannot identify a specific prospect row.

    Visual greying of other same-domain rows is handled entirely by the
    domain-grey conditional-format rule that was baked into the sheet at
    creation time (COUNTIFS formula).  That rule fires automatically once any
    row in the domain carries Lead Status = "Lead", so we must never set Lead
    on rows we don't intend to mark — doing so would inflate lead_count stats.

    Returns the number of rows updated.
    """
    lead_status = _REASON_TO_LEAD_STATUS.get(reason.lower(), "Unsubscribe")
    target      = email_or_domain.lower().strip()

    # A bare domain can't match any specific prospect row — nothing to do.
    if "@" not in target:
        return 0

    gc = _client()
    sh = gc.open_by_key(sheet_id)
    ws = sh.sheet1

    # Find column positions from the header row
    headers = ws.row_values(1)
    try:
        email_col_idx = headers.index("Recipient Email") + 1   # 1-based
        ls_col_idx    = headers.index("Lead Status")    + 1
    except ValueError:
        return 0   # sheet doesn't have the expected columns

    # Fetch the entire email column and build exact-match updates
    col_values = ws.col_values(email_col_idx)

    updates = []
    for row_num, cell_val in enumerate(col_values, start=1):
        if row_num == 1:
            continue   # header
        if str(cell_val).lower().strip() == target:
            updates.append({"range": f"{_col_letter(ls_col_idx)}{row_num}",
                             "values": [[lead_status]]})

    if updates:
        ws.batch_update(updates)

        # Ensure the CF rules are present on this sheet (retroactive for old
        # sheets created before the rules were baked into _apply_sheet_formatting).
        # - row-grey:       greys the marked row itself (non-LS columns) for ANY
        #                   status — Lead / Reply / Interested / Unsubscribe
        # - domain-grey:    Lead only — greys all other same-domain rows
        # - sender-stripe:  keeps Sender Account cell yellow (overrides everything)
        #
        # Existing rules are read first so each rule is added at most once per
        # sheet; if the lookup fails the helpers fall back to always-add.
        existing_cf = _existing_cf_formulas(sheet_id, ws.id)
        try:
            _apply_lead_grey_cf(sheet_id, ws.id, ls_col_idx, len(headers),
                                existing_formulas=existing_cf)
        except Exception as _cf_err:
            print(f"[google_sheets] lead-grey CF apply error: {_cf_err}")
        if lead_status == "Lead":
            try:
                _apply_domain_grey_cf(
                    sheet_id, ws.id, ls_col_idx, email_col_idx, len(headers),
                    existing_formulas=existing_cf,
                )
            except Exception as _cf_err:
                print(f"[google_sheets] domain-grey CF apply error: {_cf_err}")
        sender_col_idx = (
            headers.index("Sender Account") + 1
            if "Sender Account" in headers else -1
        )
        if sender_col_idx > 0:
            try:
                _apply_sender_stripe_cf(sheet_id, ws.id, sender_col_idx,
                                        existing_formulas=existing_cf)
            except Exception as _cf_err:
                print(f"[google_sheets] sender-stripe CF apply error: {_cf_err}")

    return len(updates)


# ── Internal helpers ──────────────────────────────────────────────────────────

def _add_stats_sheet(sh, headers: list, has_chaser: bool) -> None:
    """Add a 'Stats' worksheet to an existing Google Sheet with live SUMPRODUCT formulas."""

    def _h_to_col(name: str, fallback: int = 1) -> str:
        try:
            return _col_letter(headers.index(name) + 1)
        except ValueError:
            return _col_letter(fallback)

    sd_col = _h_to_col("Sent Date", 3)
    ls_col = _h_to_col("Lead Status", 14)
    ss_col = _h_to_col("Send Status", 1)

    SD   = f"'Outreach List'!${sd_col}$2:${sd_col}$10000"
    LS   = f"'Outreach List'!${ls_col}$2:${ls_col}$10000"
    SS   = f"'Outreach List'!${ss_col}:${ss_col}"
    SENT = f"IFERROR(DATEVALUE({SD}),{SD})"

    if has_chaser:
        cd_col = _h_to_col("Chaser Date", 13)
        cs_col = _h_to_col("Chaser Sent?", 12)
        CD          = f"'Outreach List'!${cd_col}$2:${cd_col}$10000"
        CS          = f"'Outreach List'!${cs_col}:${cs_col}"
        CHASER_SENT = f"IFERROR(DATEVALUE({CD}),{CD})"

    ANCHOR = f"=('Outreach List'!{sd_col}2+1)-WEEKDAY('Outreach List'!{sd_col}2+1,2)+1"

    def sp_sent(days):
        return f'=SUMPRODUCT(--({SENT}>=TODAY()-{days}),--({SENT}<TODAY()))'

    def sp_ls(days, cond):
        return f'=SUMPRODUCT(--({SENT}>=TODAY()-{days}),--({SENT}<TODAY()),(--({LS}{cond})))'

    def sp_chaser(days):
        if has_chaser:
            return f'=SUMPRODUCT(--({CHASER_SENT}>=TODAY()-{days}),--({CHASER_SENT}<TODAY()))'
        return '=0'

    N_WEEKS  = 20
    N_MONTHS = 20
    G = 7  # first data column index (1-based)

    def glet(i: int) -> str:
        return _col_letter(G + i)

    rows = [[""] * (G + N_WEEKS - 1) for _ in range(17)]

    def s(r: int, c: int, v) -> None:
        rows[r - 1][c - 1] = v

    # Row 1 summary headers
    s(1, 1, "Date"); s(1, 2, "Last day"); s(1, 3, "Last 7 days"); s(1, 4, "Last 30 days")

    # Rows 2-9: metrics
    SUMMARY = [
        ("Contacted",    sp_sent(1),               sp_sent(7),               sp_sent(30)),
        ("Chasers",      sp_chaser(1),              sp_chaser(7),             sp_chaser(30)),
        ("Replies",      sp_ls(1, '<>""'),           sp_ls(7, '<>""'),         sp_ls(30, '<>""')),
        ("Reply rate",   "=B4/B2",                  "=C4/C2",                 "=D4/D2"),
        ("Leads",        sp_ls(1, '="Lead"'),        sp_ls(7, '="Lead"'),      sp_ls(30, '="Lead"')),
        ("Lead rate",    "=B6/B2",                  "=C6/C2",                 "=D6/D2"),
        ("Interested",   sp_ls(1, '="Interested"'),  sp_ls(7, '="Interested"'), sp_ls(30, '="Interested"')),
        ("Unsubscribes", sp_ls(1, '="Unsubscribe"'), sp_ls(7, '="Unsubscribe"'), sp_ls(30, '="Unsubscribe"')),
    ]
    for ri, (label, f_day, f_7, f_30) in enumerate(SUMMARY, 2):
        s(ri, 1, label); s(ri, 2, f_day); s(ri, 3, f_7); s(ri, 4, f_30)

    # Row 13: % contacted / % chasered
    s(13, 1, "% contacted")
    s(13, 2, f'=COUNTIF({SS},TRUE)/COUNTA({SS})')
    if has_chaser:
        s(13, 3, f'=COUNTIF({CS},TRUE)/(COUNTIF({SS},TRUE)-SUMPRODUCT(--(LEN({LS})>0)))')

    # Weekly headers (row 1, cols F+)
    s(1, 6, "Weekly")
    s(1, G, ANCHOR)
    for i in range(1, N_WEEKS):
        s(1, G + i, f"={glet(i - 1)}1+7")

    # Weekly metric rows (2-8)
    WK_LABELS = ["Sends", "Replies", "Reply rate", "Leads", "Lead rate", "Interested", "Unsubscribes"]
    for mi, m_label in enumerate(WK_LABELS):
        row = mi + 2
        s(row, 6, m_label)
        for i in range(N_WEEKS):
            cl = glet(i)
            if m_label == "Sends":
                fml = f'=IF({cl}$1="","",SUMPRODUCT(--({SD}<>""),--({SENT}>={cl}$1),--({SENT}<{cl}$1+7)))'
            elif m_label == "Replies":
                fml = f'=SUMPRODUCT(--({SENT}>={cl}$1),--({SENT}<{cl}$1+7),(--({LS}<>"")))'
            elif m_label == "Reply rate":
                fml = f"={cl}3/{cl}2"
            elif m_label == "Leads":
                fml = f'=SUMPRODUCT(--({SENT}>={cl}$1),--({SENT}<{cl}$1+7),(--({LS}="Lead")))'
            elif m_label == "Lead rate":
                fml = f"={cl}5/{cl}2"
            elif m_label == "Interested":
                fml = f'=SUMPRODUCT(--({SENT}>={cl}$1),--({SENT}<{cl}$1+7),(--({LS}="Interested")))'
            else:
                fml = f'=SUMPRODUCT(--({SENT}>={cl}$1),--({SENT}<{cl}$1+7),(--({LS}="Unsubscribe")))'
            s(row, G + i, fml)

    # Monthly headers (row 10, cols F+)
    s(10, 6, "Monthly")
    s(10, G, ANCHOR)
    for i in range(1, N_MONTHS):
        s(10, G + i, f"=EDATE({glet(i - 1)}10,1)")

    # Monthly metric rows (11-17)
    MO_LABELS = ["Sends", "Replies", "Reply rate", "Leads", "Lead rate", "Interested", "Unsubscribes"]
    for mi, m_label in enumerate(MO_LABELS):
        row = mi + 11
        s(row, 6, m_label)
        for i in range(N_MONTHS):
            cl = glet(i)
            if m_label == "Sends":
                fml = f'=SUMPRODUCT(--({SENT}>={cl}$10),--({SENT}<EDATE({cl}$10,1)))'
            elif m_label == "Replies":
                fml = f'=SUMPRODUCT(--({SENT}>={cl}$10),--({SENT}<EDATE({cl}$10,1)),(--({LS}<>"")))'
            elif m_label == "Reply rate":
                fml = f"={cl}12/{cl}11"
            elif m_label == "Leads":
                fml = f'=SUMPRODUCT(--({SENT}>={cl}$10),--({SENT}<EDATE({cl}$10,1)),(--({LS}="Lead")))'
            elif m_label == "Lead rate":
                fml = f"={cl}14/{cl}11"
            elif m_label == "Interested":
                fml = f'=SUMPRODUCT(--({SENT}>={cl}$10),--({SENT}<EDATE({cl}$10,1)),(--({LS}="Interested")))'
            else:
                fml = f'=SUMPRODUCT(--({SENT}>={cl}$10),--({SENT}<EDATE({cl}$10,1)),(--({LS}="Unsubscribe")))'
            s(row, G + i, fml)

    # Write to Google Sheet
    stats_ws = sh.add_worksheet(title="Stats", rows=20, cols=G + N_WEEKS)
    end_col   = _col_letter(G + N_WEEKS - 1)
    stats_ws.update(rows, f"A1:{end_col}17", value_input_option="USER_ENTERED")

    bold = {"textFormat": {"bold": True}}
    stats_ws.format("A1:D1", bold)
    stats_ws.format(f"F1:{end_col}1", bold)
    stats_ws.format(f"F10:{end_col}10", bold)
    stats_ws.format("A1:A13", bold)
    stats_ws.format("F2:F17", bold)

    stats_ws.format(f"G1:{end_col}1",   {"numberFormat": {"type": "DATE", "pattern": "D MMM YY"}})
    stats_ws.format(f"G10:{end_col}10", {"numberFormat": {"type": "DATE", "pattern": "MMM YY"}})

    pct_fmt = {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}}
    stats_ws.format("B5:D5", pct_fmt)   # Reply rate summary
    stats_ws.format("B7:D7", pct_fmt)   # Lead rate summary
    stats_ws.format("B13:C13", pct_fmt) # % contacted / % chasered
    for i in range(N_WEEKS):
        cl = glet(i)
        stats_ws.format(f"{cl}4:{cl}4", pct_fmt)  # weekly Reply rate
        stats_ws.format(f"{cl}6:{cl}6", pct_fmt)  # weekly Lead rate
    for i in range(N_MONTHS):
        cl = glet(i)
        stats_ws.format(f"{cl}13:{cl}13", pct_fmt)  # monthly Reply rate
        stats_ws.format(f"{cl}15:{cl}15", pct_fmt)  # monthly Lead rate

    print(f"[google_sheets] Stats sheet created (id={stats_ws.id})")


def _col_letter(n: int) -> str:
    """Convert a 1-based column index to an Excel-style letter (1→A, 26→Z, 27→AA)."""
    result = ""
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result or "A"
