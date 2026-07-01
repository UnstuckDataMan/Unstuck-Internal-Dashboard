"""
Excel output writer for merge results and send schedules.

Produces professionally formatted workbooks with:
  - Coloured, grouped headers
  - Dropdown data validation on tracking fields
  - Conditional formatting for status columns
  - Auto-filter and frozen header row
"""
from typing import List, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle


# ──────────────────────────────────────────────────────────────────────────── #
# Colour tokens
# ──────────────────────────────────────────────────────────────────────────── #
C = {
    # Section header fills
    'hdr_prospect':  '1E3A5F',   # dark navy   – prospect cols
    'hdr_sender':    '1565C0',   # blue        – sender / recipient
    'hdr_template':  '0D47A1',   # darker blue – generated copy
    'hdr_chaser':    '283593',   # indigo      – chaser
    'hdr_tracking':  '1B5E20',   # dark green  – tracking
    'hdr_schedule':  '1E3A5F',   # navy        – schedule sheet
    # Row fills
    'row_even':  'EBF5FB',
    'row_odd':   'FFFFFF',
    # Status colours (conditional formatting)
    'green_bg':  'C8E6C9',
    'red_bg':    'FFCDD2',
    'amber_bg':  'FFF9C4',
    'blue_bg':   'E3F2FD',
    'purple_bg': 'F3E5F5',
    'orange_bg': 'FFE0B2',
    'lead_bg':   'D4EDD6',   # slightly darker green — Lead status
    'reply_bg':  'FFD9B3',   # pastel orange — Reply status
    'domain_bg': 'C8F0CA',   # pastel green — domain-match highlight
}

THIN_BORDER_COLOR = 'D0D7DE'


def _thin_border():
    s = Side(style='thin', color=THIN_BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr_cell(cell, fill_hex: str, font_size: int = 9):
    cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=font_size)
    cell.fill = PatternFill('solid', fgColor=fill_hex)
    cell.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
    cell.border = _thin_border()


def _data_cell(cell, fill_hex: str, wrap: bool = False, bold: bool = False):
    cell.font = Font(name='Calibri', size=9, bold=bold)
    cell.fill = PatternFill('solid', fgColor=fill_hex)
    cell.alignment = Alignment(vertical='top', wrap_text=wrap)
    cell.border = _thin_border()


def _add_dv(ws, formula: str, col_letter: str, max_row: int):
    dv = DataValidation(type='list', formula1=formula, showDropDown=False)
    dv.sqref = f"{col_letter}2:{col_letter}{max_row}"
    ws.add_data_validation(dv)


def _cf_rule(formula: str, fill_hex: str) -> Rule:
    """Build a formula-based CF rule with a properly registered dxf fill.

    FormulaRule/CellIsRule both fail to reliably register the PatternFill in
    the workbook's dxf styles block. Using Rule + DifferentialStyle directly
    is the only approach that works across all openpyxl versions.
    Setting both fgColor and bgColor is required by Excel for solid fills
    in differential formatting contexts.
    """
    fill = PatternFill(patternType='solid', fgColor=fill_hex, bgColor=fill_hex)
    dxf  = DifferentialStyle(fill=fill)
    return Rule(type='expression', dxf=dxf, formula=[formula])


def _add_cf_equal(ws, data_range: str, match_value: str, fill_hex: str):
    start_cell = data_range.split(':')[0]
    col = ''.join(c for c in start_cell if c.isalpha())
    row = ''.join(c for c in start_cell if c.isdigit())
    ws.conditional_formatting.add(
        data_range,
        _cf_rule(f'${col}{row}="{match_value}"', fill_hex),
    )


# ──────────────────────────────────────────────────────────────────────────── #
# MERGE OUTPUT
# ──────────────────────────────────────────────────────────────────────────── #

def write_merge_output(
    output_path: str,
    prospect_headers: List[str],
    merged_rows: List[Dict],
    has_chaser: bool,
    email_column: str = '',
    has_schedule: bool = False,
):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Outreach List'
    ws.freeze_panes = 'A2'

    # ── Column layout ──────────────────────────────────────────────────────
    # Section 1: send status checkbox — always first
    sec_status = ['Send Status']
    # Section 2: schedule — Sent Date always sits right after Send Time so its
    #             column letter is fixed every time a schedule is produced
    sec_schedule = ['Send Time', 'Sent Date'] if has_schedule else []
    # Section 3: sender / recipient
    sec_routing = ['Sender Account', 'First Name', 'Recipient Email']
    # Section 4: generated copy
    sec_template = ['Subject Line', 'Email Body', 'A/B Variant']
    if has_schedule:
        sec_template += ['Chaser Send Time']
    if has_chaser:
        # Chaser Sent? and Chaser Date always follow Chaser Body in a fixed block
        sec_template += ['Chaser Body', 'Chaser Sent?', 'Chaser Date']
    # Section 5: tracking
    sec_tracking = ['Lead Status', 'Notes']
    # Section 6: grey divider + all original prospect columns from the uploaded file
    _DIV = '__divider__'
    sec_divider  = [_DIV]
    sec_prospect = list(prospect_headers) if prospect_headers else []

    all_cols = (sec_status + sec_schedule + sec_routing + sec_template
                + sec_tracking + sec_divider + sec_prospect)

    # Map header → colour token
    color_map: Dict[str, str] = {}
    for h in sec_status:
        color_map[h] = C['hdr_tracking']
    for h in sec_schedule:
        color_map[h] = C['hdr_tracking']
    for h in sec_routing:
        color_map[h] = C['hdr_sender']
    for h in ['Subject Line', 'Email Body', 'A/B Variant']:
        color_map[h] = C['hdr_template']
    for h in ['Chaser Body', 'Chaser Send Time', 'Chaser Sent?', 'Chaser Date']:
        color_map[h] = C['hdr_chaser']
    color_map['Sent Date'] = C['hdr_tracking']
    for h in sec_tracking:
        color_map[h] = C['hdr_tracking']
    for h in sec_prospect:
        color_map[h] = C['hdr_prospect']

    # Column widths
    col_widths = {
        'Send Status': 10,
        'Send Time': 12,      'Sent Date': 14,
        'Sender Account': 30, 'First Name': 18, 'Recipient Email': 34,
        'Subject Line': 44,   'Email Body': 64,
        'A/B Variant': 12,    'Chaser Send Time': 14,
        'Chaser Body': 64,    'Chaser Sent?': 12, 'Chaser Date': 14,
        'Lead Status': 18,    'Notes': 32,
        _DIV: 3,
    }

    # ── Write header row ───────────────────────────────────────────────────
    for ci, header in enumerate(all_cols, 1):
        if header == _DIV:
            # Write the marker string so google_sheets.py can locate this
            # column by name without needing fragile fill-colour detection.
            # Font colour matches the fill (BDBDBD) → visually invisible in Excel.
            cell = ws.cell(row=1, column=ci, value='__divider__')
            cell.fill   = PatternFill('solid', fgColor='BDBDBD')
            cell.font   = Font(name='Calibri', color='BDBDBD', size=9)
            cell.border = _thin_border()
            ws.column_dimensions[get_column_letter(ci)].width = col_widths[_DIV]
        else:
            cell = ws.cell(row=1, column=ci, value=header)
            _hdr_cell(cell, color_map.get(header, C['hdr_prospect']))
            ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(header, 20)

    ws.row_dimensions[1].height = 20

    # ── Data validation ────────────────────────────────────────────────────
    col_map = {h: i + 1 for i, h in enumerate(all_cols)}
    total_data_rows = len(merged_rows)

    # Group rows by date so we can insert a separator after each day
    date_groups: list = []
    for row_data in merged_rows:
        d = row_data.get('__send_date__', '')
        if not date_groups or date_groups[-1][0] != d:
            date_groups.append((d, []))
        date_groups[-1][1].append(row_data)

    n_separators = len(date_groups)
    last_row = total_data_rows + n_separators + 1  # header + data + separators

    def col_let(name):
        return get_column_letter(col_map[name])

    _add_dv(ws, '"Sent"',
            col_let('Send Status'), last_row)
    _add_dv(ws, '"Lead,Reply,Unsubscribe"',
            col_let('Lead Status'), last_row)

    # ── Write data rows with end-of-day separators ─────────────────────────
    n_cols    = len(all_cols)
    wrap_cols = {'Email Body', 'Chaser Body', 'Notes'}
    ri = 2
    prev_sender = None

    for _date_val, day_rows in date_groups:
        for row_data in day_rows:
            current_sender = row_data.get('__sender_account__', '')
            is_first_of_sender = (current_sender != prev_sender)
            prev_sender = current_sender

            fill_hex = 'FFFDE7' if is_first_of_sender else 'FFFFFF'

            # Resolve first name from prospect data — try common column name variants
            _fn_key = next(
                (k for k in row_data
                 if k.lower().replace(' ', '').replace('_', '') in ('firstname', 'fname', 'first')),
                None,
            )
            first_name = row_data.get(_fn_key, '') if _fn_key else ''

            row_values = {
                'Send Status': '',
                'Send Time':   row_data.get('__send_time__', ''),
                'Sent Date':   '',
                'Sender Account': row_data.get('__sender_account__', ''),
                'First Name':     first_name,
                'Recipient Email': row_data.get('__recipient_email__', ''),
                'Subject Line': row_data.get('__subject_line__', ''),
                'Email Body':   row_data.get('__email_body__', ''),
                'A/B Variant':  row_data.get('__template_variant__', ''),
                'Chaser Send Time': row_data.get('__chaser_send_time__', ''),
                'Chaser Body':  row_data.get('__chaser_body__', ''),
                'Chaser Sent?': '',
                'Chaser Date':  '',
                'Lead Status':  '',
                'Notes':        '',
            }

            for ci, header in enumerate(all_cols, 1):
                if header == _DIV:
                    cell = ws.cell(row=ri, column=ci, value='')
                    cell.fill   = PatternFill('solid', fgColor='EEEEEE')
                    cell.border = _thin_border()
                else:
                    # For original prospect columns fall back to raw row_data
                    value = row_values.get(header, row_data.get(header, ''))
                    cell  = ws.cell(row=ri, column=ci, value=value)
                    _data_cell(cell, fill_hex, wrap=(header in wrap_cols))

            ws.row_dimensions[ri].height = 15
            ri += 1

        # ── End-of-day separator ───────────────────────────────────────────
        ws.merge_cells(f'A{ri}:{get_column_letter(n_cols)}{ri}')
        sep_cell = ws.cell(row=ri, column=1, value='No More Emails For Today.')
        sep_cell.font = Font(name='Calibri', bold=True, size=9, color='5D4037')
        sep_cell.fill = PatternFill('solid', fgColor='FFF3E0')
        sep_cell.alignment = Alignment(horizontal='center', vertical='center')
        sep_cell.border = _thin_border()
        ws.row_dimensions[ri].height = 15
        prev_sender = None  # reset so first row of next day gets yellow stripe
        ri += 1

    # ── Conditional formatting ─────────────────────────────────────────────
    def cf_range(name):
        return f"{col_let(name)}2:{col_let(name)}{last_row}"

    # Cell-level rules added first — in Excel, rules added earlier get higher
    # priority, so these will correctly override the whole-row Sent highlight.
    _add_cf_equal(ws, cf_range('Lead Status'), 'Lead',        C['lead_bg'])
    _add_cf_equal(ws, cf_range('Lead Status'), 'Reply',       C['reply_bg'])
    _add_cf_equal(ws, cf_range('Lead Status'), 'Unsubscribe', C['red_bg'])

    # Domain-match highlight: when any other row has Lead Status = "Lead" and
    # shares the same email domain, highlight that Recipient Email cell.
    e_col  = col_let('Recipient Email')
    ls_col = col_let('Lead Status')
    domain_formula = (
        f'=IFERROR(SUMPRODUCT('
        f'(${ls_col}$2:${ls_col}${last_row}="Lead")'
        f'*IFERROR((MID(${e_col}$2:${e_col}${last_row},FIND("@",${e_col}$2:${e_col}${last_row})+1,255)'
        f'=MID({e_col}2,FIND("@",{e_col}2)+1,255))*1,0)'
        f'*(ROW(${e_col}$2:${e_col}${last_row})<>ROW({e_col}2))'
        f')>0,FALSE)'
    )
    fill = PatternFill(patternType='solid', fgColor=C['domain_bg'], bgColor=C['domain_bg'])
    dxf  = DifferentialStyle(fill=fill)
    domain_rule = Rule(type='expression', dxf=dxf, formula=[domain_formula])
    ws.conditional_formatting.add(cf_range('Recipient Email'), domain_rule)

    # Whole-row Sent highlight added last so it has lowest priority and does
    # not override the cell-level Lead Status / domain-match rules above.
    full_range = f"A2:{get_column_letter(len(all_cols))}{last_row}"
    ws.conditional_formatting.add(full_range, _cf_rule('$A2="Sent"', 'E8F5E9'))

    # ── Auto-filter ────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_cols))}1"

    # ── Summary sheet ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    _write_merge_summary(ws2, merged_rows, has_chaser)

    # ── Stats sheet ────────────────────────────────────────────────────────
    _write_stats_sheet(wb, col_map, has_chaser)

    wb.save(output_path)


def _write_merge_summary(ws, merged_rows: List[Dict], has_chaser: bool):
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 22

    t = ws.cell(row=1, column=1, value='Campaign Merge Summary')
    t.font = Font(name='Calibri', bold=True, size=14, color='1E3A5F')

    senders = {r.get('__sender_account__', '') for r in merged_rows}
    variants = {r.get('__template_variant__', '') for r in merged_rows}

    stats = [
        ('Total Prospects Merged',  len(merged_rows)),
        ('Unique Sender Accounts',  len(senders)),
        ('Template Variants',        len(variants)),
        ('Chaser Templates Included', 'Yes' if has_chaser else 'No'),
    ]

    for ri, (label, value) in enumerate(stats, 3):
        lc = ws.cell(row=ri, column=1, value=label)
        lc.font = Font(name='Calibri', bold=True, size=10)
        lc.fill = PatternFill('solid', fgColor='EBF5FB')
        vc = ws.cell(row=ri, column=2, value=value)
        vc.font = Font(name='Calibri', size=10)

    # Sender breakdown
    sender_counts: Dict[str, int] = {}
    for r in merged_rows:
        s = r.get('__sender_account__', 'Unknown')
        sender_counts[s] = sender_counts.get(s, 0) + 1

    row = len(stats) + 5
    ws.cell(row=row, column=1, value='Sender Distribution').font = Font(
        bold=True, size=11, color='1E3A5F')
    row += 1

    for sender, count in sorted(sender_counts.items()):
        ws.cell(row=row, column=1, value=sender)
        ws.cell(row=row, column=2, value=count)
        row += 1


def _write_stats_sheet(wb: Workbook, col_map: Dict[str, int], has_chaser: bool):
    """Add a 'Stats' worksheet with live COUNTIFS formulas referencing Outreach List."""
    ws = wb.create_sheet('Stats')

    # ── Column references into Outreach List ──────────────────────────────
    def _ref(col_name: str, fallback: int) -> str:
        letter = get_column_letter(col_map.get(col_name, fallback))
        return f"'Outreach List'!{letter}2:{letter}10000"

    SD = _ref('Sent Date', 3)
    CD = _ref('Chaser Date', 13) if has_chaser else None
    LS = _ref('Lead Status', 14)
    EM = _ref('Recipient Email', 6)

    T  = 'TODAY()'
    WK = f'({T}-WEEKDAY({T},2)+1)'  # Monday of current week

    # Formula: first Monday of the campaign (based on earliest Sent Date)
    sd_col = get_column_letter(col_map.get('Sent Date', 3))
    _SD_FULL = f"'Outreach List'!{sd_col}2:{sd_col}10000"
    WEEK0 = (
        f'MINIFS({_SD_FULL},{_SD_FULL},"<>")'
        f'-WEEKDAY(MINIFS({_SD_FULL},{_SD_FULL},"<>"),2)+1'
    )

    # ── Style helpers ─────────────────────────────────────────────────────
    NAVY  = '1E3A5F'
    GREEN = '1B5E20'
    DBLUE = '0D47A1'

    def _h(cell, bg: str):
        cell.font      = Font(name='Calibri', bold=True, color='FFFFFF', size=9)
        cell.fill      = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = _thin_border()

    def _lbl(cell):
        cell.font      = Font(name='Calibri', bold=True, size=9, color='1A1A2E')
        cell.fill      = PatternFill('solid', fgColor='F0F4F8')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border    = _thin_border()

    def _val(cell, pct: bool = False):
        cell.font          = Font(name='Calibri', size=9)
        cell.fill          = PatternFill('solid', fgColor='FFFFFF')
        cell.alignment     = Alignment(horizontal='center', vertical='center')
        cell.border        = _thin_border()
        if pct:
            cell.number_format = '0.00%'

    # ── Summary panel: columns A–D, rows 1–12 ────────────────────────────
    for ci, (label, bg) in enumerate(
        [('Metric', NAVY), ('Today', GREEN), ('This Week', GREEN), ('All Time', GREEN)], 1
    ):
        _h(ws.cell(row=1, column=ci, value=label), bg)

    # Rows 2–9: Contacted / Chasers / Replies / Reply rate / Leads / Lead rate / Interested / Unsubscribes
    METRICS = [
        ('Contacted',
            f'=COUNTIFS({SD},">="&{T},{SD},"<"&{T}+1)',
            f'=COUNTIFS({SD},">="&{WK},{SD},"<="&{T})',
            f'=COUNTIF({SD},"<>")',
            False),
        ('Chasers',
            f'=COUNTIFS({CD},">="&{T},{CD},"<"&{T}+1)' if has_chaser else '=0',
            f'=COUNTIFS({CD},">="&{WK},{CD},"<="&{T})' if has_chaser else '=0',
            f'=COUNTIF({CD},"<>")'                       if has_chaser else '=0',
            False),
        ('Replies',
            f'=COUNTIFS({SD},">="&{T},{SD},"<"&{T}+1,{LS},"Reply")',
            f'=COUNTIFS({SD},">="&{WK},{SD},"<="&{T},{LS},"Reply")',
            f'=COUNTIF({LS},"Reply")',
            False),
        ('Reply rate',    '=IFERROR(B4/B2,"")', '=IFERROR(C4/C2,"")', '=IFERROR(D4/D2,"")', True),
        ('Leads',
            f'=COUNTIFS({SD},">="&{T},{SD},"<"&{T}+1,{LS},"Lead")',
            f'=COUNTIFS({SD},">="&{WK},{SD},"<="&{T},{LS},"Lead")',
            f'=COUNTIF({LS},"Lead")',
            False),
        ('Lead rate',     '=IFERROR(B6/B2,"")', '=IFERROR(C6/C2,"")', '=IFERROR(D6/D2,"")', True),
        ('Interested',
            f'=COUNTIFS({SD},">="&{T},{SD},"<"&{T}+1,{LS},"Interested")',
            f'=COUNTIFS({SD},">="&{WK},{SD},"<="&{T},{LS},"Interested")',
            f'=COUNTIF({LS},"Interested")',
            False),
        ('Unsubscribes',
            f'=COUNTIFS({SD},">="&{T},{SD},"<"&{T}+1,{LS},"Unsubscribe")',
            f'=COUNTIFS({SD},">="&{WK},{SD},"<="&{T},{LS},"Unsubscribe")',
            f'=COUNTIF({LS},"Unsubscribe")',
            False),
    ]

    for ri, (label, f_today, f_week, f_all, is_pct) in enumerate(METRICS, 2):
        _lbl(ws.cell(row=ri, column=1, value=label))
        for ci, fml in enumerate([f_today, f_week, f_all], 2):
            _val(ws.cell(row=ri, column=ci, value=fml), pct=is_pct)

    # Row 11: % Contacted (all-time sent / total prospects with an email)
    _lbl(ws.cell(row=11, column=1, value='% Contacted'))
    _val(ws.cell(row=11, column=2, value=f'=IFERROR(D2/COUNTIF({EM},"*@*"),"")'), pct=True)
    ws.cell(row=11, column=3, value='')
    ws.cell(row=11, column=4, value='')

    # Row 12: % Chasered (chasers sent / contacted)
    _lbl(ws.cell(row=12, column=1, value='% Chasered'))
    _val(ws.cell(row=12, column=2, value='=IFERROR(D3/D2,"")' if has_chaser else '=0'), pct=True)

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 3   # visual gap

    # ── Weekly breakdown: column F = labels, G+ = one column per week ────
    N_WEEKS  = 20
    N_MONTHS = 20
    F_COL = 6   # column F
    G_COL = 7   # first data column

    _h(ws.cell(row=1, column=F_COL, value='Weekly'), NAVY)

    # Week date headers (row 1): G1 = first Monday, H1 = G1+7, ...
    ws.cell(row=1, column=G_COL, value=f'={WEEK0}').number_format = 'D MMM YY'
    _h(ws.cell(row=1, column=G_COL), DBLUE)
    ws.column_dimensions[get_column_letter(G_COL)].width = 11
    for i in range(1, N_WEEKS):
        col_letter = get_column_letter(G_COL + i)
        c = ws.cell(row=1, column=G_COL + i, value=f'={get_column_letter(G_COL + i - 1)}1+7')
        c.number_format = 'D MMM YY'
        _h(c, DBLUE)
        ws.column_dimensions[col_letter].width = 11

    # Weekly metric rows (2–8)
    WK_METRICS = [
        ('Sends',        False),
        ('Replies',      False),
        ('Reply rate',   True),
        ('Leads',        False),
        ('Lead rate',    True),
        ('Interested',   False),
        ('Unsubscribes', False),
    ]
    for mi, (m_label, is_pct) in enumerate(WK_METRICS):
        row = mi + 2
        _lbl(ws.cell(row=row, column=F_COL, value=m_label))
        for i in range(N_WEEKS):
            cl = get_column_letter(G_COL + i)
            if m_label == 'Sends':
                fml = f'=COUNTIFS({SD},">="&{cl}1,{SD},"<"&{cl}1+7)'
            elif m_label == 'Replies':
                fml = f'=COUNTIFS({SD},">="&{cl}1,{SD},"<"&{cl}1+7,{LS},"Reply")'
            elif m_label == 'Reply rate':
                fml = f'=IFERROR({cl}3/{cl}2,"")'
            elif m_label == 'Leads':
                fml = f'=COUNTIFS({SD},">="&{cl}1,{SD},"<"&{cl}1+7,{LS},"Lead")'
            elif m_label == 'Lead rate':
                fml = f'=IFERROR({cl}5/{cl}2,"")'
            elif m_label == 'Interested':
                fml = f'=COUNTIFS({SD},">="&{cl}1,{SD},"<"&{cl}1+7,{LS},"Interested")'
            else:  # Unsubscribes
                fml = f'=COUNTIFS({SD},">="&{cl}1,{SD},"<"&{cl}1+7,{LS},"Unsubscribe")'
            _val(ws.cell(row=row, column=G_COL + i, value=fml), pct=is_pct)

    # ── Monthly breakdown: same F column, rows 10–17 ──────────────────────
    _h(ws.cell(row=10, column=F_COL, value='Monthly'), NAVY)

    # Month date headers (row 10): G10 = same as G1, H10 = EDATE(G10,1), ...
    first_month_cell = get_column_letter(G_COL) + '1'
    c = ws.cell(row=10, column=G_COL, value=f'={first_month_cell}')
    c.number_format = 'MMM YY'
    _h(c, DBLUE)
    for i in range(1, N_MONTHS):
        cl = get_column_letter(G_COL + i)
        prev_cl = get_column_letter(G_COL + i - 1)
        c = ws.cell(row=10, column=G_COL + i, value=f'=EDATE({prev_cl}10,1)')
        c.number_format = 'MMM YY'
        _h(c, DBLUE)

    # Monthly metric rows (11–17)
    MO_METRICS = [
        ('Sends',        False),
        ('Replies',      False),
        ('Reply rate',   True),
        ('Leads',        False),
        ('Lead rate',    True),
        ('Interested',   False),
        ('Unsubscribes', False),
    ]
    for mi, (m_label, is_pct) in enumerate(MO_METRICS):
        row = mi + 11
        _lbl(ws.cell(row=row, column=F_COL, value=m_label))
        for i in range(N_MONTHS):
            cl = get_column_letter(G_COL + i)
            # Next month start: use next column's row-10 date, or far future for last col
            next_date = f'{get_column_letter(G_COL + i + 1)}10' if i < N_MONTHS - 1 else 'DATE(2099,1,1)'
            if m_label == 'Sends':
                fml = f'=COUNTIFS({SD},">="&{cl}10,{SD},"<"&{next_date})'
            elif m_label == 'Replies':
                fml = f'=COUNTIFS({SD},">="&{cl}10,{SD},"<"&{next_date},{LS},"Reply")'
            elif m_label == 'Reply rate':
                fml = f'=IFERROR({cl}12/{cl}11,"")'
            elif m_label == 'Leads':
                fml = f'=COUNTIFS({SD},">="&{cl}10,{SD},"<"&{next_date},{LS},"Lead")'
            elif m_label == 'Lead rate':
                fml = f'=IFERROR({cl}14/{cl}11,"")'
            elif m_label == 'Interested':
                fml = f'=COUNTIFS({SD},">="&{cl}10,{SD},"<"&{next_date},{LS},"Interested")'
            else:  # Unsubscribes
                fml = f'=COUNTIFS({SD},">="&{cl}10,{SD},"<"&{next_date},{LS},"Unsubscribe")'
            _val(ws.cell(row=row, column=G_COL + i, value=fml), pct=is_pct)

    ws.column_dimensions[get_column_letter(F_COL)].width = 14
    ws.freeze_panes = 'B2'
