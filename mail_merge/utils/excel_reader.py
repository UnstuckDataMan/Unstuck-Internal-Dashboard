"""
Excel prospect file parser.
"""
from openpyxl import load_workbook
from typing import List, Dict, Tuple


def parse_prospect_file(filepath: str) -> Tuple[List[str], List[Dict[str, str]], int]:
    """
    Load an Excel file and return:
      - headers: list of column header strings (from row 1)
      - all_rows: list of dicts mapping header → cell value (as strings)
      - total_rows: number of data rows
    """
    wb = load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)

    # First row is headers
    header_row = next(rows_iter, None)
    if header_row is None:
        raise ValueError("The spreadsheet appears to be empty.")

    headers = []
    for cell in header_row:
        if cell is not None and str(cell).strip():
            headers.append(str(cell).strip())

    if not headers:
        raise ValueError("No column headers found in the first row.")

    all_rows: List[Dict[str, str]] = []

    for raw_row in rows_iter:
        # Skip entirely blank rows
        if all(c is None for c in raw_row):
            continue

        row_dict: Dict[str, str] = {}
        for i, header in enumerate(headers):
            cell_value = raw_row[i] if i < len(raw_row) else None
            if cell_value is None:
                row_dict[header] = ''
            else:
                row_dict[header] = str(cell_value).strip()
        all_rows.append(row_dict)

    wb.close()
    return headers, all_rows, len(all_rows)
