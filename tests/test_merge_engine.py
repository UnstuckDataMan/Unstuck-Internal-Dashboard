"""Offline tests for the merge engine, scheduler, and Excel writer."""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "mail_merge"))

from utils.merge import (
    _expand_inline_variants, perform_merge, reassign_templates, validate_templates,
)
from utils.scheduler import generate_schedule
from utils.excel_writer import write_merge_output

HEADERS = ["First Name", "Company", "Email"]
SUBJ    = ["Hi {{first_name}}", "Hello {{first_name}}", "Hey {{first_name}}"]
BODY    = ["Body one {{company}}", "Body two {{company}}"]
SENDERS = ["s1@unstuck.com", "s2@unstuck.com"]


def make_rows(n):
    return [{"First Name": f"P{i}", "Company": f"C{i}", "Email": f"p{i}@dom{i}.com"}
            for i in range(n)]


# ── Template validation ────────────────────────────────────────────────────────

def test_validate_accepts_fuzzy_placeholder_names():
    assert validate_templates(["{{First Name}} {{first_name}} {{FIRSTNAME}}"], HEADERS) == []


def test_validate_rejects_unknown_placeholder():
    errors = validate_templates(["{{nonexistent}}"], HEADERS)
    assert len(errors) == 1
    assert "nonexistent" in errors[0]
    assert "First Name" in errors[0]   # lists available columns


# ── Inline {a|b} variants ─────────────────────────────────────────────────────

def test_inline_variants_rotate_with_copy_index():
    t = "{Alpha|Beta} and {One|Two|Three}"
    assert _expand_inline_variants(t, 0) == "Alpha and Two"
    assert _expand_inline_variants(t, 1) == "Beta and Three"
    assert _expand_inline_variants(t, 2) == "Alpha and One"


def test_inline_variants_leave_placeholders_intact():
    out = _expand_inline_variants("{Hi|Hey} {{first_name}}", 0)
    assert out == "Hi {{first_name}}"


# ── perform_merge contract ────────────────────────────────────────────────────

def test_perform_merge_sets_recipient_and_chaser_only():
    rows = make_rows(4)
    merged = perform_merge(rows, HEADERS, SUBJ, BODY, "Bump {{first_name}}",
                           SENDERS, "[MISSING]", "Email")
    assert [r["__recipient_email__"] for r in merged] == [r["Email"] for r in rows]
    assert merged[0]["__chaser_body__"] == "Bump P0"
    # Subject/body/variant/sender are assigned later by schedule + reassign
    for r in merged:
        assert "__subject_line__" not in r
        assert "__sender_account__" not in r


def test_perform_merge_autodetects_email_column():
    merged = perform_merge(make_rows(2), HEADERS, SUBJ, BODY, "", SENDERS)
    assert merged[0]["__recipient_email__"] == "p0@dom0.com"


def test_perform_merge_requires_templates_and_senders():
    with pytest.raises(ValueError):
        perform_merge(make_rows(1), HEADERS, [], BODY, "", SENDERS)
    with pytest.raises(ValueError):
        perform_merge(make_rows(1), HEADERS, SUBJ, [], "", SENDERS)
    with pytest.raises(ValueError):
        perform_merge(make_rows(1), HEADERS, SUBJ, BODY, "", [])


def test_reassign_templates_rotates_in_final_order():
    rows = [dict(r) for r in make_rows(6)]
    reassign_templates(rows, SUBJ, BODY, HEADERS, "[MISSING]")
    assert [r["__template_variant__"] for r in rows] == [
        "S1/B1", "S2/B2", "S3/B1", "S1/B2", "S2/B1", "S3/B2",
    ]
    assert rows[0]["__subject_line__"] == "Hi P0"
    assert rows[1]["__email_body__"] == "Body two C1"


def test_missing_value_substituted_for_empty_cell():
    rows = [{"First Name": "", "Company": "ACME", "Email": "a@b.com"}]
    reassign_templates(rows, ["Hi {{first_name}}"], ["B {{company}}"], HEADERS, "[MISSING]")
    assert rows[0]["__subject_line__"] == "Hi [MISSING]"


# ── Scheduler invariants ──────────────────────────────────────────────────────

def test_schedule_covers_every_prospect_exactly_once():
    sched = generate_schedule(75, SENDERS + ["s3@unstuck.com"],
                              campaign_seed="seed1", max_per_sender_per_day=10)
    ids = [s["prospect_id"] for s in sched]
    assert sorted(ids) == list(range(1, 76))


def test_schedule_skips_weekends_and_caps_full_days():
    sched = generate_schedule(75, SENDERS + ["s3@unstuck.com"],
                              campaign_seed="seed1", max_per_sender_per_day=10)
    per_day: dict[str, int] = {}
    for s in sched:
        d = date.fromisoformat(s["date"])
        assert d.weekday() < 5, f"weekend send on {d}"
        per_day[s["date"]] = per_day.get(s["date"], 0) + 1
    days = sorted(per_day)
    # Full days carry exactly nominal × senders; only the last may be partial
    for d in days[:-1]:
        assert per_day[d] == 30, per_day
    assert sum(per_day.values()) == 75


def test_schedule_is_deterministic_for_same_seed():
    a = generate_schedule(20, SENDERS, campaign_seed="fixed", max_per_sender_per_day=10)
    b = generate_schedule(20, SENDERS, campaign_seed="fixed", max_per_sender_per_day=10)
    assert a == b


def test_schedule_enforces_min_gap_per_sender():
    sched = generate_schedule(30, SENDERS, campaign_seed="gapseed",
                              max_per_sender_per_day=15)
    by_sender_day: dict[tuple, list[str]] = {}
    for s in sched:
        by_sender_day.setdefault((s["sender"], s["date"]), []).append(s["send_time"])
    for times in by_sender_day.values():
        mins = sorted(int(t[:2]) * 60 + int(t[3:]) for t in times)
        gaps = [b - a for a, b in zip(mins, mins[1:])]
        assert all(g >= 3 for g in gaps), gaps


def test_schedule_rejects_short_window():
    with pytest.raises(ValueError):
        generate_schedule(5, SENDERS, window_start="09:00", window_end="12:00")


def test_send_instant_orders_across_midnight():
    """Sends must order chronologically even when the sender's clock rolls past
    midnight relative to the recipient's window.

    With a PT recipient window and a SAST sender (9h ahead), 08:30-15:30 PT maps
    to 17:30-00:30 SAST, so a day's last sends read "00:13"/"01:07". Sorting on
    that text put them at the TOP of the block instead of the end; 'send_instant'
    is the absolute instant and sorts correctly.
    """
    senders = ["a@x.com", "b@x.com"]
    sched = generate_schedule(
        40, senders, campaign_seed="tzseed", max_per_sender_per_day=10,
        window_start="08:30", window_end="15:30",
        recipient_tz="America/Vancouver", sender_tz="Africa/Johannesburg",
    )

    after_midnight = [s for s in sched if s["send_time"] < "08:00"]
    assert after_midnight, "fixture must actually cross midnight, or this proves nothing"

    order = {e: i for i, e in enumerate(senders)}
    by_instant = sorted(sched, key=lambda s: (s["date"], order[s["sender"]], s["send_instant"]))

    blocks: dict[tuple, list] = {}
    for s in by_instant:
        blocks.setdefault((s["date"], s["sender"]), []).append(s)

    for key, rows in blocks.items():
        instants = [r["send_instant"] for r in rows]
        assert instants == sorted(instants), f"{key} not chronological"
        # The block must END after midnight, never start there.
        if any(r["send_time"] < "08:00" for r in rows):
            assert rows[-1]["send_time"] < "08:00", \
                f"{key}: post-midnight send should be last, got {[r['send_time'] for r in rows]}"
            assert rows[0]["send_time"] > "12:00", \
                f"{key}: block should open with an afternoon-SAST time"

    # And confirm the old text sort really was broken, so this test has teeth.
    by_text = sorted(sched, key=lambda s: (s["date"], order[s["sender"]], s["send_time"]))
    assert by_text != by_instant, "text sort should differ — fixture no longer exercises the bug"


# ── Excel writer ──────────────────────────────────────────────────────────────

def _build_workbook(tmp_path, has_chaser=True):
    rows = [dict(r) for r in make_rows(6)]
    # Two date groups, two senders
    for i, r in enumerate(rows):
        r["__send_date__"]       = "2026-06-15" if i < 3 else "2026-06-16"
        r["__send_time__"]       = f"09:{10 + i:02d}"
        r["__sender_account__"]  = SENDERS[i % 2]
        r["__recipient_email__"] = r["Email"]
        r["__chaser_send_time__"] = "09:30"
        if has_chaser:
            r["__chaser_body__"] = f"Bump P{i}"
    reassign_templates(rows, SUBJ, BODY, HEADERS, "[MISSING]")
    out = tmp_path / "merge.xlsx"
    write_merge_output(str(out), HEADERS, rows, has_chaser, "Email", has_schedule=True)
    import openpyxl
    return openpyxl.load_workbook(str(out))


def test_excel_layout_and_separators(tmp_path):
    wb = _build_workbook(tmp_path)
    ws = wb["Outreach List"]
    headers = [c.value for c in ws[1]]
    for col in ["Send Status", "Send Time", "Sender Account", "Recipient Email",
                "Subject Line", "Email Body", "A/B Variant", "Chaser Send Time",
                "Chaser Body", "Lead Status", "Notes", "__divider__"]:
        assert col in headers, f"missing column {col}"
    # Original prospect columns come after the divider ("First Name" exists
    # twice by design: once as a routing column, once as the raw prospect col)
    assert headers.index("__divider__") < headers.index("Company")
    assert headers.count("First Name") == 2
    # One separator row per send date (2 dates)
    seps = [r for r in ws.iter_rows(min_row=2)
            if r[0].value == "No More Emails For Today."]
    assert len(seps) == 2
    assert "Summary" in wb.sheetnames


def test_excel_conditional_formats_and_validation(tmp_path):
    wb = _build_workbook(tmp_path)
    ws = wb["Outreach List"]
    dv_formulas = {dv.formula1 for dv in ws.data_validations.dataValidation}
    assert '"Sent"' in dv_formulas
    assert '"Lead,Reply,Unsubscribe"' in dv_formulas
    all_rules = []
    for rng in ws.conditional_formatting:
        all_rules.extend(rng.rules)
    formulas = [f for rule in all_rules for f in (rule.formula or [])]
    # Lead/Reply/Unsubscribe cell rules + domain-match + whole-row Sent
    assert any('="Lead"' in f for f in formulas)
    assert any('="Reply"' in f for f in formulas)
    assert any('="Unsubscribe"' in f for f in formulas)
    assert any("SUMPRODUCT" in f for f in formulas)       # domain-match rule
    assert any('$A2="Sent"' in f for f in formulas)        # whole-row sent
